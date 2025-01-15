from openpyxl import Workbook
import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
import openpyxl.utils
from openpyxl.worksheet.table import Table, TableStyleInfo
from django.http import FileResponse, HttpResponse
from io import BytesIO
from Apps.User.models.UserModel import User
from Apps.Tasks.models.TasksModels import Task
from Apps.Projects.models.ProjectsModels import Project
from Apps.WorkTeams.models.WorkTeamModel import WorkTeam

cell_border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'), top=Side(border_style='thin'), bottom=Side(border_style='thin'))
cell_alignment = Alignment(horizontal='center', vertical='center')
cell_font = Font(bold=True, size=12, color="FFFFFF")
patternFill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
# Estilo de la tabla
style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False, showLastColumn=False, showRowStripes=False, showColumnStripes=False)

# Ajustar ancho de columnas
def ajustar_ancho_columna(a,b, ws):
    for col in range(a,b):
        max_length = 0
        column = ws.iter_cols(min_col=col, max_col = col)
        for cell in next(column):
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
            adjusted_width = (max_length + 5)
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = adjusted_width
            
    return ws

def generar_xls_usuarios():
    # Crear un libro de trabajo y hoja
    wb = Workbook()
    ws = wb.active
    ws.title = "Listado Usuarios"
    
    headers = ["Nombre", "Apellido", "Rut", "Correo", "Telefono", "Cargo"]
    
    # Establecer los encabezados
    for col_num, header in enumerate(headers, start=3):
        cell = ws.cell(row=3, column=col_num, value=header)
        cell.font = cell_font
        header_fill = patternFill
        cell.fill = header_fill
        # Borde de las celdas
        cell.border = cell_border
        # Alineación de los encabezados
        cell.alignment = cell_alignment
    # Obtener los usuarios
    users = User.objects.all()
    # Escribir los datos de los usuarios
    fila = 4
    for user in users:
        ws.cell(row=fila, column=3, value=user.user_name)
        ws.cell(row=fila, column=4, value=user.user_last_name)
        ws.cell(row=fila, column=5, value=user.user_rut)
        ws.cell(row=fila, column=6, value=user.user_email)
        ws.cell(row=fila, column=7, value=user.user_phone)
        ws.cell(row=fila, column=8, value=user.user_role)
        fila += 1
        
    # Definir el rango de la tabla
    tabla_range = f"C3:H{fila-1}"
    # Crear la tabla
    table = Table(displayName="UsuariosTabla", ref=tabla_range)
    table.tableStyleInfo = style
    # Agregar la tabla a la hoja
    ws.add_table(table)
    # Ajustar el ancho de las columnas
    ws = ajustar_ancho_columna(3,9,ws)

    # Devolver el archivo como una respuesta HTTP
    response = HttpResponse(content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=usuarios.xlsx'
    wb.save(response)
    return response

def generar_xlsx_tareas(fecha_inicio, fecha_fin):
    wb = Workbook()
    ws = wb.active
    ws.title = "Listado Tareas"
    
    headers = ["Nombre", "Descripción", "Usuario", "Proyecto", "Fecha inicio", "estado"]
    
    for col_num, headers in enumerate(headers, start=3):
        cell = ws.cell(row=3, column=col_num, value=headers)
        cell.font = cell_font
        header_fill = patternFill
        cell.fill = header_fill
        # Borde de la celda
        cell.border = cell_border
        # Alineación de encabezados
        cell.alignment = cell_alignment
        
    # obtener todas las tareas
    tasks = Task.objects.filter(task_date_create__gte=fecha_inicio, task_date_create__lte=fecha_fin)
    # Escribir las tareas
    fila = 4
    for task in tasks:
        user = str(task.task_user)
        project = str(task.task_project)
        ws.cell(row=fila, column=3, value=task.task_name)
        ws.cell(row=fila, column=4, value=task.task_description)
        ws.cell(row=fila, column=5, value=user)
        ws.cell(row=fila, column=6, value=project)
        ws.cell(row=fila, column=7, value=task.task_date_create.strftime('%d/%m/%Y'))
        ws.cell(row=fila, column=8, value=task.task_status)
        fila += 1
        
    # Rango de la tabla
    table_range = f"C3:H{fila-1}"
    
    # Crear tabla
    table = Table(displayName="TareasTabla", ref=table_range)
    table.tableStyleInfo = style   
    # Agregar tabla a la hoja
    ws.add_table(table)
    # Ajustar el ancho de las columnas
    ws = ajustar_ancho_columna(3,9,ws)
    
    response = HttpResponse(content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=tareas.xlsx'
    wb.save(response)    
    # Devolver el archivo como respuesta HTTP
    return response

def generar_xlsx_proyectos(fecha_inicio, fecha_fin):
    wb = Workbook()
    ws = wb.active
    ws.title = "Listado Proyectos"
    headers = ["Nombre", "Descripción", "Inversión", "Fecha Inicio"]
    
    for col_num, header in enumerate(headers, start=3):
        cell = ws.cell(row=3, column=col_num, value=header)
        cell.font = cell_font
        header_fill = patternFill
        cell.fill = header_fill
        cell.border = cell_border
        cell.alignment = cell_alignment
    projects = Project.objects.filter(project_start_date__gte=fecha_inicio,project_start_date__lte=fecha_fin)
    
    fila = 4
    for project in projects:
        ws.cell(row=fila, column=3, value=project.project_name)
        ws.cell(row=fila, column=4, value=project.project_description)
        ws.cell(row=fila, column=5, value=project.project_budget)
        ws.cell(row=fila, column=6, value=project.project_start_date.strftime('%d/%m/%Y'))
        fila += 1
        
    table_range = f"C3:F{fila-1}"
    table = Table(displayName="ProyectosTabla", ref=table_range)
    table.tableStyleInfo = style
    ws.add_table(table)
    ws = ajustar_ancho_columna(3,7,ws)
    
    response = HttpResponse(content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=proyectos.xlsx'
    wb.save(response)
    return response

def generar_xlsx_equipos(fecha_inicio, fecha_fin):
    wb = Workbook()
    ws = wb.active
    ws.title = "Listado Equipos de trabajo"
    headers = ["Nombre", "Fecha Inicio", "Proyecto Actual"]
    
    for col_num, header in enumerate(headers, start=3):
        cell = ws.cell(row=3, column=col_num, value=header)
        cell.font = cell_font
        header_fill = patternFill
        cell.fill = header_fill
        cell.border = cell_border
        cell.alignment = cell_alignment
    workteams = WorkTeam.objects.filter(team_date_create__gte=fecha_inicio, team_date_create__lte=fecha_fin)
    
    fila = 4
    for workteam in workteams:
        project = str(workteam.team_project)
        ws.cell(row=fila, column=3, value=workteam.team_name)
        ws.cell(row=fila, column=4, value=workteam.team_date_create.strftime('%d/%m/%Y'))
        ws.cell(row=fila, column=5, value=project)
        fila += 1
        
    table_range = f"C3:E{fila-1}"
    table = Table(displayName="EquiposTabla", ref=table_range)
    table.tableStyleInfo = style
    ws.add_table(table)
    ws = ajustar_ancho_columna(3,6,ws)
    
    response = HttpResponse(content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=equipos.xlsx'
    wb.save(response)
    return response